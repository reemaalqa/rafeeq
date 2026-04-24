#!/usr/bin/env python3
"""
SauDial: Saudi Arabic Dialects Game Localization Dataset Generator

This script generates a comprehensive dataset for video game localization
into Saudi Arabic dialects using OpenAI's GPT-4o model.

Authors: Naif Alanazi, Mohammed Al-Batineh, Hussein Abu-Rayyash
Paper: "SauDial: The Saudi Arabic Dialects Game Localization Dataset"

Overview:
---------
The script loops over every combination of (dialect, scenario, game genre,
tone, age rating), sends a prompt to GPT-4o to generate authentic Saudi
Arabic game dialogue, validates the response, deduplicates it via hashing,
and finally exports the dataset as a formatted Excel workbook.
This is the tool used to CREATE the SauDial dataset consumed elsewhere
in this project — you only need to run it if you want to regenerate data.
"""

# ── third-party imports ───────────────────────────────────────────────────
import openai                          # OpenAI Python SDK for GPT-4o calls
import pandas as pd                    # tabular data handling
import json                            # parsing LLM JSON responses
import hashlib                         # content hashing for deduplication
import time                            # retry back-off / rate limiting
import random                          # shuffling parameter combinations
from typing import Dict, List, Any, Optional
# openpyxl is used to write nicely formatted Excel output.
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils.dataframe import dataframe_to_rows
import logging
import sys
from datetime import datetime

# Configure logging: all events are written both to stdout and to a log file
# so long-running generation runs can be inspected later.
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(levelname)s - %(message)s',
    handlers=[
        logging.FileHandler('saudial_generation.log'),
        logging.StreamHandler(sys.stdout)
    ]
)
logger = logging.getLogger(__name__)

class SauDialGenerator:
    """
    Main class for generating the SauDial dataset using OpenAI GPT-4o.
    """
    
    def __init__(self, api_key: str):
        """
        Initialize the SauDial generator with OpenAI API key.
        
        Args:
            api_key (str): OpenAI API key for GPT-4o access
        """
        # OpenAI client used for all GPT-4o chat completions.
        self.client = openai.OpenAI(api_key=api_key)
        # Hash set for deduplication — every accepted entry's content
        # is hashed so identical generations are rejected.
        self.generated_hashes = set()
        # Accumulator of all valid generated entries.
        self.data_entries = []

        # Initialise dialect linguistic knowledge and generation parameters.
        self._setup_dialect_resources()
        self._setup_generation_parameters()
        
    def _setup_dialect_resources(self):
        """
        Setup dialect-specific linguistic resources based on scholarly sources.
        """
        # For every Saudi dialect we store: a short description, key
        # phonological/lexical features, common expressions, and cultural
        # markers.  This dictionary is injected into the prompt so GPT-4o
        # can produce authentic, region-specific output.
        self.dialect_resources = {
            "Najdi": {
                "description": "Central Arabian dialect spoken in Riyadh and surrounding regions",
                "key_features": [
                    "Distinctive /q/ pronunciation as [g]",
                    "Vocabulary influenced by Bedouin heritage",
                    "Formal register closer to MSA in urban areas"
                ],
                "common_expressions": [
                    "إيش (eish) - what",
                    "وين (wein) - where", 
                    "كيفك (kifak) - how are you",
                    "الحين (al-heen) - now"
                ],
                "cultural_markers": [
                    "Desert heritage references",
                    "Traditional hospitality expressions",
                    "Camel and desert terminology"
                ]
            },
            "Hijazi": {
                "description": "Western Saudi dialect spoken in Mecca, Medina, and Jeddah",
                "key_features": [
                    "Urban cosmopolitan influences",
                    "Egyptian and Levantine borrowings",
                    "Pilgrimage-related vocabulary"
                ],
                "common_expressions": [
                    "إيه (eh) - what",
                    "فين (fein) - where",
                    "إزيك (izzayak) - how are you",
                    "دلوقت (delwa'ti) - now"
                ],
                "cultural_markers": [
                    "Pilgrimage and religious references",
                    "Urban coastal lifestyle",
                    "Trade and commerce terminology"
                ]
            },
            "Eastern": {
                "description": "Eastern Province dialect with Gulf influences",
                "key_features": [
                    "Persian Gulf linguistic features",
                    "Oil industry terminology",
                    "Maritime vocabulary"
                ],
                "common_expressions": [
                    "شنو (shinu) - what",
                    "وين (wein) - where",
                    "شلونك (shlonak) - how are you",
                    "الحين (al-heen) - now"
                ],
                "cultural_markers": [
                    "Oil industry references",
                    "Pearl diving heritage",
                    "Gulf maritime culture"
                ]
            },
            "Janoubi": {
                "description": "Southern dialect including Asir and Najran regions",
                "key_features": [
                    "Mountain highland influences",
                    "Agricultural terminology",
                    "Yemeni linguistic connections"
                ],
                "common_expressions": [
                    "إيش (eish) - what",
                    "وين (wein) - where",
                    "كيف حالك (kif halak) - how are you",
                    "الآن (al-aan) - now"
                ],
                "cultural_markers": [
                    "Mountain and agricultural life",
                    "Traditional architecture terms",
                    "Highland cultural practices"
                ]
            }
        }
    
    def _setup_generation_parameters(self):
        """
        Define all generation parameters as specified in the methodology.
        """
        # The four Saudi dialects covered by the dataset.
        self.dialects = ["Najdi", "Hijazi", "Eastern", "Janoubi"]

        # Game-world scenarios covering a wide thematic range.
        self.scenarios = [
            "Epic Battle", "Marketplace Haggling", "Family Gathering", "Desert Adventure",
            "Urban Exploration", "Ancient Ruins", "Royal Court", "Merchant Caravan",
            "Coastal Journey", "Mountain Expedition", "Tribal Council", "Festival Celebration",
            "Archaeological Site", "Trading Post", "Nomadic Camp", "City Gates",
            "Palace Intrigue", "Desert Racing", "Treasure Hunt", "Cultural Exchange",
            "Religious Pilgrimage", "Diplomatic Mission", "Village Life", "Oasis Discovery",
            "Historical Battle", "Jeddah Tower Escape Room"
        ]
        
        # Game genres used as one of the parameter-combination axes.
        self.game_genres = [
            "RPG", "Adventure", "Platformer", "Action", "Simulation",
            "Strategy", "Puzzle", "Educational"
        ]

        # Emotional tones — shape the register/feel of the generated dialogue.
        self.tones = [
            "Serious", "Humorous", "Whimsical", "Mysterious", "Inspirational",
            "Dark humor", "Contemplative", "Fearful", "Excited", "Reflective"
        ]
        
        # PEGI-style age ratings for content appropriateness.
        self.age_ratings = ["+3", "+7", "12+", "+16", "18+"]

        # Human-readable description of each tone, fed to the LLM so
        # it knows exactly what emotional register to produce.
        self.tone_definitions = {
            "Serious": "Formal language with minimal colloquialisms, authoritative tone",
            "Humorous": "Integrates jokes, puns, and playful expressions",
            "Whimsical": "Imaginative, light-hearted phrasing with creative elements",
            "Mysterious": "Ambiguous language, suspenseful and intriguing tone",
            "Inspirational": "Motivational messages, encouraging and uplifting",
            "Dark humor": "Wit combined with macabre or morbid elements",
            "Contemplative": "Introspective, thought-provoking philosophical tone",
            "Fearful": "Communicates anxiety, tension, or dread",
            "Excited": "Energetic, enthusiastic, high-energy delivery",
            "Reflective": "Reminiscent, emotional retrospection, nostalgic"
        }
    
    def _generate_content_hash(self, content: Dict[str, Any]) -> str:
        """
        Generate a unique hash for content deduplication.
        
        Args:
            content (Dict): Content dictionary to hash
            
        Returns:
            str: MD5 hash of the content
        """
        # Use MD5 over (english_text + dialect_translation) as the
        # deduplication key — this catches near-duplicates that only
        # differ in metadata fields.
        content_str = f"{content.get('english_text', '')}{content.get('dialect_translation', '')}"
        return hashlib.md5(content_str.encode('utf-8')).hexdigest()
    
    def _create_dynamic_prompt(self, dialect: str, scenario: str, game_genre: str, 
                             tone: str, age_rating: str) -> str:
        """
        Create a dynamic prompt for the GPT-4o model based on parameters.
        
        Args:
            dialect (str): Target Saudi dialect
            scenario (str): Game scenario context
            game_genre (str): Video game genre
            tone (str): Emotional tone
            age_rating (str): Age rating category
            
        Returns:
            str: Formatted prompt for the LLM
        """
        # Fetch the dialect description + tone definition for inclusion
        # in the prompt context.
        dialect_info = self.dialect_resources[dialect]
        tone_definition = self.tone_definitions[tone]

        # Multi-section prompt instructing GPT-4o to return a JSON
        # object with exactly the fields we need for the dataset.
        prompt = f"""
You are an expert in Saudi Arabic dialects and video game localization. Generate authentic game dialogue content for the {dialect} dialect.

**Context Parameters:**
- Dialect: {dialect} ({dialect_info['description']})
- Scenario: {scenario}
- Game Genre: {game_genre}
- Tone: {tone} ({tone_definition})
- Age Rating: {age_rating}

**Dialect Features to Include:**
{chr(10).join([f"- {feature}" for feature in dialect_info['key_features']])}

**Cultural Markers:**
{chr(10).join([f"- {marker}" for marker in dialect_info['cultural_markers']])}

**Task:** Create a game dialogue entry that includes:

1. **English Text**: Original game dialogue (10-25 words) appropriate for the scenario and tone
2. **MSA Translation**: Modern Standard Arabic translation maintaining meaning and formality
3. **Dialect Translation**: Authentic {dialect} dialect version using regional vocabulary, pronunciation patterns, and cultural expressions
4. **Context and Rating**: Explain the cultural context, age appropriateness, and scenario relevance (50-100 words)
5. **Dialect Notes**: Linguistic notes about dialect-specific vocabulary, pronunciation, or cultural adaptations used (30-60 words)

**Requirements:**
- Use authentic {dialect} expressions and vocabulary
- Maintain the {tone} emotional tone throughout
- Ensure age-appropriate content for {age_rating} rating
- Include cultural references relevant to {scenario}
- Make dialogue suitable for {game_genre} games

Please format your response as JSON:
{{
    "english_text": "...",
    "msa_translation": "...",
    "dialect_translation": "...",
    "context_and_rating": "...",
    "dialect_notes": "..."
}}
"""
        return prompt
    
    def _assess_localization_difficulty(self, entry: Dict[str, Any]) -> int:
        """
        Assess localization difficulty on a 1-5 scale based on multiple criteria.
        
        Args:
            entry (Dict): Generated content entry
            
        Returns:
            int: Difficulty rating (1-5)
        """
        # Heuristic difficulty scoring — starts at 1 and gains a point
        # for each complicating factor present in the entry.
        difficulty_score = 1

        # Check for cultural specificity
        cultural_terms = ["traditional", "heritage", "pilgrimage", "tribal", "royal", "ancient"]
        if any(term in entry.get('english_text', '').lower() for term in cultural_terms):
            difficulty_score += 1
            
        # Check for pragmatic complexity
        pragmatic_markers = ["humor", "whimsical", "mysterious", "dark humor"]
        if entry.get('tone', '') in pragmatic_markers:
            difficulty_score += 1
            
        # Check for technical terminology
        technical_terms = ["battle", "strategy", "puzzle", "simulation"]
        if any(term in entry.get('game_genre', '').lower() for term in technical_terms):
            difficulty_score += 1
            
        # Check for dialect-specific vocabulary richness
        dialect_notes = entry.get('dialect_notes', '')
        if len(dialect_notes.split()) > 40:
            difficulty_score += 1
            
        return min(difficulty_score, 5)
    
    def _calculate_word_counts(self, entry: Dict[str, Any]) -> Dict[str, int]:
        """
        Calculate word counts for different text fields.
        
        Args:
            entry (Dict): Content entry
            
        Returns:
            Dict: Word counts for each text field
        """
        return {
            'english_words': len(entry.get('english_text', '').split()),
            'msa_words': len(entry.get('msa_translation', '').split()),
            'dialect_words': len(entry.get('dialect_translation', '').split())
        }
    
    def generate_content(self, dialect: str, scenario: str, game_genre: str, 
                        tone: str, age_rating: str, max_retries: int = 3) -> Optional[Dict[str, Any]]:
        """
        Generate localization content using GPT-4o with error handling and retry logic.
        
        Args:
            dialect (str): Target dialect
            scenario (str): Game scenario
            game_genre (str): Game genre
            tone (str): Emotional tone
            age_rating (str): Age rating
            max_retries (int): Maximum retry attempts
            
        Returns:
            Optional[Dict]: Generated content or None if failed
        """
        # Build the per-combination prompt injected into the chat API.
        prompt = self._create_dynamic_prompt(dialect, scenario, game_genre, tone, age_rating)

        # Retry loop: the LLM sometimes returns malformed JSON or
        # duplicated content — we try up to max_retries times.
        for attempt in range(max_retries):
            try:
                response = self.client.chat.completions.create(
                    model="gpt-4o",
                    messages=[
                        {"role": "system", "content": "You are an expert linguist specializing in Saudi Arabic dialects and video game localization."},
                        {"role": "user", "content": prompt}
                    ],
                    temperature=0.7,
                    max_tokens=800
                )
                
                content_text = response.choices[0].message.content.strip()

                # Try to parse JSON response
                try:
                    # Strip markdown code fences if the model wrapped the JSON.
                    if content_text.startswith('```json'):
                        content_text = content_text.replace('```json', '').replace('```', '').strip()

                    content = json.loads(content_text)

                    # Validate that every expected field is present.
                    required_fields = ['english_text', 'msa_translation', 'dialect_translation',
                                     'context_and_rating', 'dialect_notes']
                    
                    if all(field in content for field in required_fields):
                        # Add metadata fields (these come from the caller,
                        # not from the LLM) so the entry has full context.
                        content.update({
                            'dialect': dialect,
                            'scenario': scenario,
                            'game_genre': game_genre,
                            'tone': tone,
                            'age_rating': age_rating,
                            'localization_difficulty': self._assess_localization_difficulty(content),
                            'in_game_context': f"Player engaged in {scenario.lower()} within a {game_genre.lower()} game environment"
                        })
                        
                        # Add word counts for each text field.
                        content.update(self._calculate_word_counts(content))

                        # Reject duplicates via content hashing.
                        content_hash = self._generate_content_hash(content)
                        if content_hash not in self.generated_hashes:
                            self.generated_hashes.add(content_hash)
                            return content
                        else:
                            logger.warning(f"Duplicate content detected, retrying... (Attempt {attempt + 1})")
                            continue
                    else:
                        logger.warning(f"Missing required fields in response (Attempt {attempt + 1})")
                        
                except json.JSONDecodeError as e:
                    logger.warning(f"JSON parsing error: {e} (Attempt {attempt + 1})")
                    
            except Exception as e:
                # Any network / API failure — wait longer each retry.
                logger.error(f"API call failed: {e} (Attempt {attempt + 1})")
                if attempt < max_retries - 1:
                    time.sleep(2 ** attempt)  # Exponential backoff
                    
        logger.error(f"Failed to generate content after {max_retries} attempts")
        return None
    
    def collect_data(self, samples_per_combination: int = 1, 
                    max_total_samples: Optional[int] = None) -> List[Dict[str, Any]]:
        """
        Collect data for all parameter combinations with progress tracking.
        
        Args:
            samples_per_combination (int): Number of samples per parameter combination
            max_total_samples (Optional[int]): Maximum total samples to generate
            
        Returns:
            List[Dict]: Generated data entries
        """
        # Total number of unique parameter combinations possible.
        total_combinations = (len(self.dialects) * len(self.scenarios) *
                            len(self.game_genres) * len(self.tones) *
                            len(self.age_ratings))
        
        total_planned = total_combinations * samples_per_combination
        if max_total_samples:
            total_planned = min(total_planned, max_total_samples)
            
        logger.info(f"Starting data collection for {total_planned} entries across {total_combinations} combinations")
        
        generated_count = 0

        # Build the Cartesian product of all five parameter axes.
        combinations = []
        for dialect in self.dialects:
            for scenario in self.scenarios:
                for genre in self.game_genres:
                    for tone in self.tones:
                        for rating in self.age_ratings:
                            combinations.append((dialect, scenario, genre, tone, rating))
        
        # Shuffle so early-stopped runs still cover a diverse sample.
        random.shuffle(combinations)

        # Main generation loop over every parameter combination.
        for combination in combinations:
            if max_total_samples and generated_count >= max_total_samples:
                break
                
            dialect, scenario, genre, tone, rating = combination
            
            for sample_num in range(samples_per_combination):
                if max_total_samples and generated_count >= max_total_samples:
                    break
                    
                logger.info(f"Generating entry {generated_count + 1}/{total_planned}: "
                          f"{dialect}, {scenario}, {genre}, {tone}, {rating}")
                
                content = self.generate_content(dialect, scenario, genre, tone, rating)
                
                if content:
                    self.data_entries.append(content)
                    generated_count += 1
                    
                    # Progress indicator
                    if generated_count % 10 == 0:
                        logger.info(f"Progress: {generated_count}/{total_planned} entries generated")
                        
                # Simple rate limiting — one second between API calls
                # to stay well under OpenAI's rate limits.
                time.sleep(1)
                
        logger.info(f"Data collection completed: {len(self.data_entries)} entries generated")
        return self.data_entries
    
    def create_excel_dataset(self, filename: str = "SauDial_Dataset.xlsx"):
        """
        Create formatted Excel file with the generated dataset.
        
        Args:
            filename (str): Output Excel filename
        """
        if not self.data_entries:
            logger.error("No data entries to export")
            return

        # Build a pandas DataFrame from the accumulated list of dicts.
        df = pd.DataFrame(self.data_entries)
        
        # Reorder columns to match paper specification
        column_order = [
            'dialect', 'scenario', 'game_genre', 'tone', 'age_rating',
            'english_text', 'msa_translation', 'dialect_translation',
            'context_and_rating', 'dialect_notes', 'localization_difficulty',
            'in_game_context', 'english_words', 'msa_words', 'dialect_words'
        ]
        
        df = df.reindex(columns=column_order)
        
        # Rename columns for presentation
        df.columns = [
            'Dialect', 'Scenario', 'Game Type', 'Tone', 'Age Rating',
            'English Text', 'MSA Translation', 'Dialect Translation',
            'Context and Rating', 'Dialect Notes', 'Localization Difficulty',
            'In-Game Context', 'English Words', 'MSA Words', 'Dialect Words'
        ]
        
        # Create a new Excel workbook with a single sheet for the data.
        wb = Workbook()
        ws = wb.active
        ws.title = "SauDial Dataset"

        # Define a bold white header font on a dark-blue fill.
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")

        # Append every row of the DataFrame (including the header).
        for r in dataframe_to_rows(df, index=False, header=True):
            ws.append(r)
            
        # Format header row
        for cell in ws[1]:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center", vertical="center")
            
        # Auto-adjust column widths based on the longest cell value.
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
                    
            adjusted_width = min(max_length + 2, 50)  # Cap at 50 characters
            ws.column_dimensions[column_letter].width = adjusted_width
            
        # Highlight dialect-specific words in red (as mentioned in paper).
        # This makes the dialect column visually pop in the spreadsheet.
        red_font = Font(color="FF0000")
        dialect_col = None
        
        for idx, cell in enumerate(ws[1], 1):
            if cell.value == "Dialect Translation":
                dialect_col = idx
                break
                
        if dialect_col:
            for row in range(2, ws.max_row + 1):
                ws.cell(row=row, column=dialect_col).font = red_font
                
        # Add a second worksheet with run metadata (date, counts, etc.).
        meta_ws = wb.create_sheet("Metadata")
        meta_ws.append(["SauDial Dataset Metadata"])
        meta_ws.append(["Generated on:", datetime.now().strftime("%Y-%m-%d %H:%M:%S")])
        meta_ws.append(["Total entries:", len(self.data_entries)])
        meta_ws.append(["Dialects covered:", ", ".join(self.dialects)])
        meta_ws.append(["Game genres:", ", ".join(self.game_genres)])
        meta_ws.append(["Emotional tones:", ", ".join(self.tones)])
        meta_ws.append(["Age ratings:", ", ".join(self.age_ratings)])
        
        # Persist the workbook to disk.
        wb.save(filename)
        logger.info(f"Dataset exported to {filename}")

        # Also print summary statistics for the generated dataset.
        self._generate_summary_statistics(df)
    
    def _generate_summary_statistics(self, df: pd.DataFrame):
        """
        Generate and log summary statistics for the dataset.
        
        Args:
            df (pd.DataFrame): Dataset DataFrame
        """
        logger.info("=== DATASET SUMMARY STATISTICS ===")
        logger.info(f"Total entries: {len(df)}")
        logger.info(f"Dialects: {df['Dialect'].value_counts().to_dict()}")
        logger.info(f"Game genres: {df['Game Type'].value_counts().to_dict()}")
        logger.info(f"Tones: {df['Tone'].value_counts().to_dict()}")
        logger.info(f"Age ratings: {df['Age Rating'].value_counts().to_dict()}")
        logger.info(f"Localization difficulty distribution: {df['Localization Difficulty'].value_counts().sort_index().to_dict()}")
        logger.info(f"Average English words: {df['English Words'].mean():.1f}")
        logger.info(f"Average MSA words: {df['MSA Words'].mean():.1f}")
        logger.info(f"Average dialect words: {df['Dialect Words'].mean():.1f}")


def main():
    """
    Main function to run the SauDial dataset generation.
    """
    print("SauDial: Saudi Arabic Dialects Game Localization Dataset Generator")
    print("=" * 60)

    # Prompt user for their OpenAI API key (never hard-coded).
    api_key = input("Enter your OpenAI API key: ").strip()
    if not api_key:
        print("Error: OpenAI API key is required")
        return
        
    # Instantiate the main generator class.
    try:
        generator = SauDialGenerator(api_key)
        logger.info("SauDial generator initialized successfully")
    except Exception as e:
        logger.error(f"Failed to initialize generator: {e}")
        return
        
    # Ask user how much data to generate (with sensible defaults).
    try:
        max_samples = int(input("Enter maximum number of samples to generate (default: 100): ") or "100")
        samples_per_combo = int(input("Enter samples per parameter combination (default: 1): ") or "1")
    except ValueError:
        logger.error("Invalid input for sample counts")
        return
        
    # Run the full data-collection pipeline and export results.
    try:
        logger.info("Starting dataset generation...")
        generator.collect_data(
            samples_per_combination=samples_per_combo,
            max_total_samples=max_samples
        )
        
        # Export to Excel
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        filename = f"SauDial_Dataset_{timestamp}.xlsx"
        generator.create_excel_dataset(filename)
        
        print(f"\nDataset generation completed successfully!")
        print(f"Generated {len(generator.data_entries)} entries")
        print(f"Dataset saved as: {filename}")
        
    except KeyboardInterrupt:
        # User pressed Ctrl+C — save whatever we have so far as a
        # partial dataset rather than losing the work.
        logger.info("Generation interrupted by user")
        if generator.data_entries:
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            filename = f"SauDial_Dataset_Partial_{timestamp}.xlsx"
            generator.create_excel_dataset(filename)
            print(f"Partial dataset saved as: {filename}")
            
    except Exception as e:
        logger.error(f"Error during generation: {e}")


if __name__ == "__main__":
    main()
